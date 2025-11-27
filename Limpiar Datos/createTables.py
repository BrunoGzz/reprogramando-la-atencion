import os
import csv
from openpyxl import Workbook

base_folder = "./all_data"

for subcarpeta in os.listdir(base_folder):
    ruta_subcarpeta = os.path.join(base_folder, subcarpeta)
    if os.path.isdir(ruta_subcarpeta):
        # Leer todos los CSV y organizar datos
        csv_files = sorted([f for f in os.listdir(ruta_subcarpeta) if f.endswith(".csv")])
        all_data = []
        titles = []
        max_length = 0
        grupos = {}

        for archivo in csv_files:
            ruta_csv = os.path.join(ruta_subcarpeta, archivo)
            titles.append(archivo[:5])  # título de columna
            grupo_clave = archivo[:2]
            if grupo_clave not in grupos:
                grupos[grupo_clave] = []
            grupos[grupo_clave].append((archivo, ruta_csv))

            # Leer columna Attention
            valores_attention = []
            with open(ruta_csv, newline='', encoding='utf-8') as f:
                lector = csv.DictReader(f)
                for fila in lector:
                    valores_attention.append(fila.get("Attention", ""))
            all_data.append(valores_attention)
            max_length = max(max_length, len(valores_attention))

        # Crear libro Excel
        wb = Workbook()
        ws_main = wb.active
        ws_main.title = "Attention"

        # Escribir timestamp relativo
        ws_main.cell(row=1, column=1, value="Timestamp")
        for i in range(1, max_length + 1):
            ws_main.cell(row=i + 1, column=1, value=i)

        # Escribir datos de Attention
        for col_idx, (title, valores) in enumerate(zip(titles, all_data), start=2):
            ws_main.cell(row=1, column=col_idx, value=title)
            for row_idx, valor in enumerate(valores, start=2):
                ws_main.cell(row=row_idx, column=col_idx, value=int(valor))

        # Crear hojas por grupo de dos caracteres
        for clave, lista_archivos in grupos.items():
            ws = wb.create_sheet(title=clave)
            # Orden alfabético por nombre de archivo
            lista_archivos = sorted(lista_archivos, key=lambda x: x[0])
            max_len_grupo = 0
            for _, ruta_csv in lista_archivos:
                with open(ruta_csv, newline='', encoding='utf-8') as f:
                    lector = csv.DictReader(f)
                    n = sum(1 for _ in lector)
                    if n > max_len_grupo:
                        max_len_grupo = n

            # Escribir timestamp relativo
            ws.cell(row=1, column=1, value="Timestamp")
            for i in range(1, max_len_grupo + 1):
                ws.cell(row=i + 1, column=1, value=i)

            # Escribir columnas de Attention
            for col_idx, (archivo, ruta_csv) in enumerate(lista_archivos, start=2):
                ws.cell(row=1, column=col_idx, value=archivo[:5])
                valores_attention = []
                with open(ruta_csv, newline='', encoding='utf-8') as f:
                    lector = csv.DictReader(f)
                    for fila in lector:
                        valores_attention.append(fila.get("Attention", ""))
                for row_idx, valor in enumerate(valores_attention, start=2):
                    ws.cell(row=row_idx, column=col_idx, value=int(valor))

        # Guardar archivo Excel
        excel_path = os.path.join(base_folder, f"{subcarpeta}.xlsx")
        wb.save(excel_path)
        print(f"Archivo creado: {excel_path}")
